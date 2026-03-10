"""
NTCPEvaluator — Unified performance evaluation for all NTCP model tiers (v1.1.0).

Usage:
    evaluator = NTCPEvaluator(n_bootstrap=1000, cv_n_splits=5, random_state=42)
    metrics = evaluator.evaluate(y_true, y_pred_apparent, model_name='LKB_Probit',
                                 tier='T1A', organ='Parotid', y_pred_cv=None)
"""

from dataclasses import dataclass
from typing import Optional, List

import numpy as np
import pandas as pd
from sklearn.metrics import roc_auc_score, brier_score_loss
from sklearn.calibration import calibration_curve


@dataclass
class ModelMetrics:
    """Standardized metrics container for any NTCP model."""

    model_name: str
    tier: str                          # 'T1A', 'T1B', 'T2', 'T3', 'T4'
    organ: str
    n: int
    n_events: int
    event_rate: float

    # Discrimination
    apparent_auc: float                # AUC on full data (always computed)
    apparent_auc_ci_l: float           # 95% bootstrap CI lower
    apparent_auc_ci_u: float           # 95% bootstrap CI upper
    cv_auc: Optional[float] = None     # Cross-validated AUC (None for fixed-param models)
    cv_auc_std: Optional[float] = None
    loo_auc: Optional[float] = None    # Leave-one-out AUC
    cv_strategy: str = 'N/A'           # 'LOO', '5-fold', or 'N/A' (fixed params)
    overfitting_gap: Optional[float] = None  # apparent_auc - cv_auc

    # AUC type label (critical for manuscript table)
    auc_type_label: str = 'apparent'   # 'apparent', 'LOO-CV', '5-fold-CV', 'N/A (fixed)'
    honest_auc: Optional[float] = None  # Best honest estimate: cv_auc if available, else apparent

    # Calibration
    brier_score: Optional[float] = None
    ece: Optional[float] = None        # Expected Calibration Error
    mce: Optional[float] = None        # Maximum Calibration Error
    calibration_slope: Optional[float] = None
    calibration_intercept: Optional[float] = None

    # Model properties
    n_parameters: Optional[int] = None
    epv: Optional[float] = None        # Events Per Variable
    n_features: Optional[int] = None

    # QA flags
    overfitting_flag: bool = False     # Gap > threshold OR clearly poor CV performance
    boundary_flag: bool = False        # Parameter at optimization boundary
    leakage_flag: bool = False         # Detected data leakage
    epv_flag: bool = False             # EPV below minimum

    # Additional
    bootstrap_n: int = 1000
    notes: str = ''


class NTCPEvaluator:
    """
    Unified evaluator for all NTCP model tiers.

    Implements:
    - Bootstrap 95% CI for apparent AUC (all models)
    - CV-AUC where cross-validated predictions are available
    - Brier score (all models)
    - ECE and MCE (all models)
    - Overfitting gap = apparent - cv_auc
    - Overfitting flag: gap > 0.10 OR (cv_auc < 0.55 AND apparent > 0.65)
    """

    def __init__(
        self,
        n_bootstrap: int = 1000,
        cv_n_splits: int = 5,
        loo_threshold: int = 100,
        random_state: int = 42,
        overfitting_gap_threshold: float = 0.10,
    ):
        self.n_bootstrap = n_bootstrap
        self.cv_n_splits = cv_n_splits
        self.loo_threshold = loo_threshold
        self.random_state = random_state
        self.overfitting_gap_threshold = overfitting_gap_threshold

    def _bootstrap_auc(self, y_true: np.ndarray, y_pred: np.ndarray) -> (float, float):
        """Bootstrap 95% CI for AUC."""
        rng = np.random.RandomState(self.random_state)
        y_true = np.asarray(y_true)
        y_pred = np.asarray(y_pred)
        boot_aucs: List[float] = []

        for _ in range(self.n_bootstrap):
            idx = rng.choice(len(y_true), size=len(y_true), replace=True)
            try:
                boot_aucs.append(roc_auc_score(y_true[idx], y_pred[idx]))
            except Exception:
                continue

        if not boot_aucs:
            return np.nan, np.nan
        return float(np.percentile(boot_aucs, 2.5)), float(np.percentile(boot_aucs, 97.5))

    def _calibration_metrics(self, y_true: np.ndarray, y_pred: np.ndarray, n_bins: int = 10):
        """Compute ECE, MCE, and calibration slope/intercept."""
        try:
            frac_pos, mean_pred = calibration_curve(y_true, y_pred, n_bins=n_bins)
            # Histogram for weights
            hist, _ = np.histogram(y_pred, bins=n_bins, range=(0, 1))
            bin_weights = hist / len(y_pred) if len(y_pred) > 0 else np.zeros_like(hist, dtype=float)
            # Align weights with frac_pos length
            weights = bin_weights[: len(frac_pos)]

            ece = float(np.average(np.abs(frac_pos - mean_pred), weights=weights))
            mce = float(np.max(np.abs(frac_pos - mean_pred)))

            # Simple linear regression of observed vs predicted probabilities
            if len(mean_pred) < 2:
                return np.nan, np.nan, np.nan, np.nan
            x = mean_pred
            y = frac_pos
            x_mean = np.mean(x)
            y_mean = np.mean(y)
            num = np.sum((x - x_mean) * (y - y_mean))
            den = np.sum((x - x_mean) ** 2)
            if den == 0:
                return ece, mce, np.nan, np.nan
            slope = num / den
            intercept = y_mean - slope * x_mean
            return ece, mce, float(slope), float(intercept)
        except Exception:
            return np.nan, np.nan, np.nan, np.nan

    def evaluate(
        self,
        y_true: np.ndarray,
        y_pred_apparent: np.ndarray,
        model_name: str,
        tier: str,
        organ: str,
        y_pred_cv: np.ndarray = None,
        n_parameters: int = None,
        n_features: int = None,
        cv_strategy: str = 'N/A',
        boundary_flag: bool = False,
        leakage_flag: bool = False,
        epv: float = None,
    ) -> ModelMetrics:
        """
        Evaluate a single model.

        Args:
            y_true: Binary outcomes
            y_pred_apparent: Predictions from full-data model
            model_name: Model identifier
            tier: One of 'T1A', 'T1B', 'T2', 'T3', 'T4'
            organ: Organ name
            y_pred_cv: CV predictions (if available); if None, only apparent metrics computed
            n_parameters: Number of free parameters (for EPV)
            n_features: Number of features used
            cv_strategy: 'LOO', '5-fold', or 'N/A'
            boundary_flag: True if any parameter was at optimization boundary
            leakage_flag: True if data leakage was detected
            epv: Precomputed EPV (or compute from n_events / n_features)
        """
        y_true = np.asarray(y_true).astype(int)
        y_pred_apparent = np.asarray(y_pred_apparent)

        n = len(y_true)
        n_events = int(np.sum(y_true))
        event_rate = n_events / n if n > 0 else 0.0

        # Apparent AUC + bootstrap CI
        try:
            apparent_auc = roc_auc_score(y_true, y_pred_apparent)
        except Exception:
            apparent_auc = np.nan
        ci_l, ci_u = self._bootstrap_auc(y_true, y_pred_apparent)

        # CV AUC
        cv_auc = None
        cv_auc_std = None
        loo_auc = None
        overfitting_gap = None
        if y_pred_cv is not None:
            y_pred_cv = np.asarray(y_pred_cv)
            try:
                cv_auc = float(roc_auc_score(y_true, y_pred_cv))
            except Exception:
                cv_auc = np.nan
            if cv_strategy == 'LOO':
                loo_auc = cv_auc
            # std of folds should be passed by caller where available;
            # here we keep cv_auc_std as provided, or NaN by default.
            cv_auc_std = np.nan
            if not np.isnan(apparent_auc) and not np.isnan(cv_auc):
                overfitting_gap = apparent_auc - cv_auc

        # AUC type label
        if cv_strategy == 'N/A':
            auc_type_label = 'apparent (fixed params)'
        elif cv_strategy == 'LOO':
            auc_type_label = 'LOO-CV'
        else:
            auc_type_label = '5-fold-CV'

        honest_auc = cv_auc if cv_auc is not None else apparent_auc

        # Brier, ECE, MCE
        try:
            brier = float(brier_score_loss(y_true, y_pred_apparent))
        except Exception:
            brier = np.nan
        ece, mce, cal_slope, cal_intercept = self._calibration_metrics(y_true, y_pred_apparent)

        # EPV
        if epv is None and n_features and n_features > 0:
            epv = n_events / float(n_features)
        epv_flag = (
            epv is not None
            and n_features is not None
            and n_features > 1
            and epv < 10.0
        )

        # Overfitting flag
        overfitting_flag = False
        if overfitting_gap is not None and not np.isnan(overfitting_gap):
            if cv_auc is not None and not np.isnan(cv_auc):
                overfitting_flag = (
                    overfitting_gap > self.overfitting_gap_threshold
                    or (cv_auc < 0.55 and apparent_auc > 0.65)
                )

        return ModelMetrics(
            model_name=model_name,
            tier=tier,
            organ=organ,
            n=n,
            n_events=n_events,
            event_rate=event_rate,
            apparent_auc=apparent_auc,
            apparent_auc_ci_l=ci_l,
            apparent_auc_ci_u=ci_u,
            cv_auc=cv_auc,
            cv_auc_std=cv_auc_std,
            loo_auc=loo_auc,
            cv_strategy=cv_strategy,
            overfitting_gap=overfitting_gap,
            auc_type_label=auc_type_label,
            honest_auc=honest_auc,
            brier_score=brier,
            ece=ece,
            mce=mce,
            calibration_slope=cal_slope,
            calibration_intercept=cal_intercept,
            n_parameters=n_parameters,
            epv=epv,
            n_features=n_features,
            overfitting_flag=overfitting_flag,
            boundary_flag=boundary_flag,
            leakage_flag=leakage_flag,
            epv_flag=epv_flag,
            bootstrap_n=self.n_bootstrap,
        )

    def to_dataframe(self, metrics_list: list) -> pd.DataFrame:
        """Convert list of ModelMetrics to a publication-ready DataFrame."""
        rows = []
        for m in metrics_list:
            rows.append({
                'Organ': m.organ,
                'Tier': m.tier,
                'Model': m.model_name,
                'N': m.n,
                'Events': m.n_events,
                'Event_rate_%': round(m.event_rate * 100, 1),
                'Apparent_AUC': round(m.apparent_auc, 4) if not np.isnan(m.apparent_auc) else '',
                'Apparent_AUC_95CI': (
                    f"({m.apparent_auc_ci_l:.3f}–{m.apparent_auc_ci_u:.3f})"
                    if not np.isnan(m.apparent_auc_ci_l)
                    else ''
                ),
                'CV_AUC': round(m.cv_auc, 4) if m.cv_auc is not None else 'N/A (fixed)',
                'CV_AUC_SD': round(m.cv_auc_std, 3) if m.cv_auc_std is not None else '',
                'AUC_type': m.auc_type_label,
                'Honest_AUC': round(m.honest_auc, 4) if m.honest_auc is not None else '',
                'Overfitting_Gap': (
                    round(m.overfitting_gap, 4) if m.overfitting_gap is not None else 'N/A'
                ),
                'Brier_Score': round(m.brier_score, 4) if m.brier_score is not None else '',
                'ECE': round(m.ece, 4) if m.ece is not None else '',
                'MCE': round(m.mce, 4) if m.mce is not None else '',
                'EPV': round(m.epv, 1) if m.epv is not None else '',
                'N_features': m.n_features or '',
                'Overfitting_flag': '⚠ Yes' if m.overfitting_flag else 'OK',
                'Boundary_flag': '⚠ Yes' if m.boundary_flag else 'OK',
                'Leakage_flag': '⚠ Yes' if m.leakage_flag else 'OK',
                'EPV_flag': '⚠ Yes' if m.epv_flag else 'OK',
                'Notes': m.notes,
            })
        return pd.DataFrame(rows)

    def save_performance_table(self, metrics_list: list, output_path: str):
        """Save the unified performance table as Excel with conditional formatting."""
        import openpyxl
        from openpyxl.styles import PatternFill

        df = self.to_dataframe(metrics_list)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='All_Models', index=False)

            # Pivot: Apparent vs Honest AUC comparison
            pivot_cols = [
                'Organ', 'Tier', 'Model', 'Apparent_AUC', 'CV_AUC',
                'AUC_type', 'Overfitting_Gap', 'Brier_Score', 'ECE', 'Overfitting_flag'
            ]
            df[pivot_cols].to_excel(writer, sheet_name='Performance_Summary', index=False)

            # Calibration table (only include cols that exist)
            cal_cols = [
                'Organ', 'Tier', 'Model', 'Brier_Score', 'ECE', 'MCE',
                'calibration_slope', 'calibration_intercept', 'N_features', 'EPV'
            ]
            available_cal_cols = [c for c in cal_cols if c in df.columns]
            if available_cal_cols:
                df[available_cal_cols].to_excel(writer, sheet_name='Calibration', index=False)

            # QA flags table
            flag_cols = [
                'Organ', 'Tier', 'Model', 'EPV', 'Overfitting_flag',
                'Boundary_flag', 'Leakage_flag', 'EPV_flag', 'Notes'
            ]
            df[flag_cols].to_excel(writer, sheet_name='QA_Flags', index=False)

            # Apply simple conditional formatting on Performance_Summary
            ws = writer.sheets['Performance_Summary']
            red_fill = PatternFill("solid", fgColor="FFD0D0")
            orange_fill = PatternFill("solid", fgColor="FFE8CC")
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if str(cell.value) == '⚠ Yes':
                        cell.fill = red_fill
                    # Column E is CV_AUC in our layout
                    if cell.column_letter == 'E' and cell.value not in (None, '', 'N/A (fixed)'):
                        try:
                            if float(cell.value) < 0.55:
                                cell.fill = orange_fill
                        except Exception:
                            pass

